import xlrd
import openpyxl

workbook = openpyxl.load_workbook("UserLogin.xlsx")
sheet = workbook["login"]

# get total number of rows
rowCount = sheet.max_row
print("total number of rows", rowCount)
colCount = sheet.max_column
print(colCount)

for row in range(2, rowCount + 1):
    user_name = sheet.cell(row=row, column=1).value
    password = sheet.cell(row=row, column=2).value
    username = str(user_name)
    password = str(password)
    print(username + " " + password)
